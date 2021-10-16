from django.shortcuts import render, redirect
from planificador.models import (
    Orden_compra,
    Cotizacion,
    RMC,
    Producto_proyecto,
    Producto_proyecto_cantidades,
    Producto_proveedor,
    Gastos_generales,
    Relacion_gastos,
    Proyecto,
    Correlativo_orden_compra,
    Importaciones,
)
from datetime import date
import uuid
from openpyxl import load_workbook
from RMC_Corporate.settings import EXCEL_ROOT
import json
from django.contrib.auth import get_user_model


def excel_oc(id_orden_compra):
    path = "{}/OC_FORMAT.xlsx".format(EXCEL_ROOT)
    wb = load_workbook(path)
    sheet = wb.get_sheet_by_name("OC")
    orden_compra = Orden_compra.objects.get(id=id_orden_compra)
    proveedor = orden_compra.cotizacion_hija.proveedor_asociado
    productos = orden_compra.cotizacion_hija.productos_proyecto_asociados.all()
    cotizacion = orden_compra.cotizacion_hija
    suma_productos = 0
    numero_inicial_excel = 13
    for producto in productos:
        if Producto_proveedor.objects.filter(
            proyecto=producto.producto_asociado_cantidades.proyecto, producto=proveedor
        ).exists():
            nombre_producto = Producto_proveedor.objects.filter(
                proyecto=producto.producto_asociado_cantidades.proyecto,
                producto=proveedor,
            ).nombre_proveedor
        else:
            nombre_producto = producto.producto_asociado_cantidades.proyecto.nombre
        cantidad = int(producto.cantidades.split(".")[0])
        unidad = producto.producto_asociado_cantidades.proyecto.unidad
        lista_precios = (
            producto.producto_asociado_cantidades.proyecto.lista_precios.all()
        )
        precio = lista_precios.filter(
            nombre_cotizacion=orden_compra.cotizacion_padre.nombre
        )[0].valor
        total = precio * int(cantidad)
        suma_productos += total
        sheet["B{}".format(str(numero_inicial_excel))] = nombre_producto
        sheet["C{}".format(str(numero_inicial_excel))] = cantidad
        sheet["D{}".format(str(numero_inicial_excel))] = unidad
        sheet["E{}".format(str(numero_inicial_excel))] = precio
        sheet["F{}".format(str(numero_inicial_excel))] = total
        numero_inicial_excel += 1
    for i in range(numero_inicial_excel, 23):
        if sheet["B{}".format(i)].value is not None:
            sheet["B{}".format(i)] = ""
            sheet["C{}".format(i)] = ""
            sheet["D{}".format(i)] = ""
            sheet["E{}".format(i)] = ""
            sheet["F{}".format(i)] = ""
    # TOTALES
    sheet["F4"] = orden_compra.id
    sheet["F24"] = suma_productos
    sheet["F25"] = suma_productos * 0.19
    sheet["F26"] = suma_productos * 1.19
    # PROVEEDOR
    sheet["B8"] = cotizacion.proveedor_asociado.nombre
    sheet["B9"] = cotizacion.proveedor_asociado.direccion
    sheet["B10"] = cotizacion.proveedor_asociado.contactos_asociados.all()[0].nombre
    sheet["D8"] = date.today()
    sheet["D9"] = cotizacion.proveedor_asociado.rut
    sheet["D10"] = cotizacion.proveedor_asociado.contactos_asociados.all()[0].telefono
    # CONDICIONES
    sheet["B33"] = orden_compra.condicion_entrega
    sheet["B34"] = orden_compra.condiciones_pago
    sheet["B35"] = orden_compra.forma_pago
    # EMPRESA
    sheet["A38"] = orden_compra.destino_factura.nombre
    sheet["A39"] = "GIRO: " + orden_compra.destino_factura.giro
    sheet["A40"] = "Dirección: " + orden_compra.destino_factura.direccion
    sheet["C38"] = "RUT: " + orden_compra.destino_factura.rut
    sheet["A43"] = orden_compra.observaciones
    wb.save("{}{}.xlsx".format(EXCEL_ROOT, orden_compra.id))
    orden_compra.planilla.name = "{}{}.xlsx".format(EXCEL_ROOT, orden_compra.id)
    orden_compra.save()


def subir_gasto(request, id):
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        wb = load_workbook(excel_file)
        sheet = wb.get_sheet_by_name("FORMATO REL COMPRAS")
        # RELACION DE GASTOS
        id_relacion = uuid.uuid1()
        numero_relacion = sheet["A1"].value.split(":")[1][1:]
        if not numero_relacion:
            numero_relacion = id_relacion
        fecha = sheet["A7"].value
        if not fecha:
            fecha = date.today()
        periodo_desde = sheet["U7"].value
        if not periodo_desde:
            periodo_desde = date.today()
        periodo_hasta = sheet["U8"].value
        if not periodo_hasta:
            periodo_hasta = date.today()
        rut_solicitante = sheet["A51"].value.split(":")[1][1:]
        if not rut_solicitante:
            rut_solicitante = "No hay"
        rut_autorizador = sheet["H51"].value.split(":")[1][1:]
        if not rut_autorizador:
            rut_autorizador = "No hay"
        rut_aprobador = sheet["N51"].value.split(":")[1][1:]
        if not rut_aprobador:
            rut_aprobador = "No hay"
        nueva_relacion_gasto = Relacion_gastos(
            id=id_relacion,
            numero_relacion=numero_relacion,
            fecha=fecha,
            periodo_desde=periodo_desde,
            periodo_hasta=periodo_hasta,
            rut_solicitante=rut_solicitante,
            rut_autorizador=rut_autorizador,
            rut_aprobador=rut_aprobador,
        )
        nueva_relacion_gasto.save()
        # GASTOS GENERALES
        numero_facturas = 19
        for i in range(numero_facturas, 28):
            if (
                not sheet["A{}".format(i)].value
                and not sheet["C{}".format(i)].value
                and not sheet["O{}".format(i)].value
                and not sheet["T{}".format(i)].value
            ):
                pass
            else:
                fecha = sheet["A{}".format(i)].value
                if not fecha:
                    fecha = periodo_hasta
                numero_factura = sheet["C{}".format(i)].value
                if not numero_factura:
                    numero_factura = "No hay"
                razon_social = sheet["I{}".format(i)].value
                if not razon_social:
                    razon_social = "No hay"
                detalle = sheet["O{}".format(i)].value
                if not detalle:
                    detalle = "No hay"
                monto = sheet["T{}".format(i)].value
                if not monto:
                    monto = 0
                nuevo_gasto_general = Gastos_generales(
                    id=uuid.uuid1(),
                    fecha=fecha,
                    numero_factura=numero_factura,
                    factura_o_boleta="factura",
                    razon_social=razon_social,
                    detalle=detalle,
                    monto=monto,
                )
                nuevo_gasto_general.save()
                nueva_relacion_gasto.gastos_generales.add(nuevo_gasto_general)
                nueva_relacion_gasto.save()
                if not nueva_relacion_gasto.total_factura:
                    nueva_relacion_gasto.total_factura = 0
                nueva_relacion_gasto.total_factura += monto
                nueva_relacion_gasto.save()
            numero_facturas += 1
        numero_boletas = 32
        for i in range(numero_boletas, 41):
            if (
                not sheet["A{}".format(i)].value
                and not sheet["C{}".format(i)].value
                and not sheet["O{}".format(i)].value
                and not sheet["T{}".format(i)].value
            ):
                pass
            else:
                fecha = sheet["A{}".format(i)].value
                if not fecha:
                    fecha = periodo_hasta
                numero_factura = sheet["C{}".format(i)].value
                if not numero_factura:
                    numero_factura = "No hay"
                razon_social = sheet["I{}".format(i)].value
                if not razon_social:
                    razon_social = "No hay"
                detalle = sheet["O{}".format(i)].value
                if not detalle:
                    detalle = "No hay"
                monto = sheet["T{}".format(i)].value
                if not monto:
                    monto = 0
                nuevo_gasto_general = Gastos_generales(
                    id=uuid.uuid1(),
                    fecha=fecha,
                    numero_factura=numero_factura,
                    factura_o_boleta="boleta",
                    razon_social=razon_social,
                    detalle=detalle,
                    monto=monto,
                )
                nuevo_gasto_general.save()
                nueva_relacion_gasto.gastos_generales.add(nuevo_gasto_general)
                if not nueva_relacion_gasto.total_boleta:
                    nueva_relacion_gasto.total_boleta = 0
                nueva_relacion_gasto.total_boleta += monto
                nueva_relacion_gasto.save()
            numero_boletas += 1
        proyecto = Proyecto.objects.get(id=id)
        proyecto.relacion_gastos.add(nueva_relacion_gasto)
        proyecto.save()
        return redirect("/proyectos/proyecto/{}".format(id))
    else:
        return render(request, "proyectos/subir_gasto.html")


def crear_orden(request, id):
    if request.method == "GET":
        cotizacion_padre = Cotizacion.objects.get(id=id)
        productos_proyecto = []
        for producto in cotizacion_padre.productos_asociados.all():
            producto_proyecto = Producto_proyecto.objects.get(
                producto=cotizacion_padre.proyecto_asociado, proyecto=producto
            )
            productos_proyecto.append(producto_proyecto)
        return render(
            request,
            "orden_compra/crear_orden.html",
            {"cotizacion_padre": cotizacion_padre, "productos": productos_proyecto},
        )
    else:
        id_cotizacion = request.POST["id_cotizacion"]
        date_today = date.today().year
        if Correlativo_orden_compra.objects.filter(año=date_today).exists():
            correlativo = Correlativo_orden_compra.objects.get(año=date_today)
            id_orden = correlativo.numero
            correlativo.numero = correlativo.numero + 1
            correlativo.save()
        else:
            nuevo_correlativo = Correlativo_orden_compra(año=2021, numero=2100)
            nuevo_correlativo.save()
            id_orden = nuevo_correlativo.numero
        condicion_entrega = request.POST["condicion_entrega"]
        condicion_pago = request.POST["condicion_pago"]
        forma_pago = request.POST["forma_pago"]
        factura = request.POST["empresa"]
        observaciones = request.POST["observaciones"]
        productos = request.POST.getlist("id_producto")
        cotizacion_padre = Cotizacion.objects.get(id=id_cotizacion)
        nuevo_nombre = cotizacion_padre.nombre + " - " + str(id_orden)
        fecha_envio = date.today()
        nueva_cotizacion = Cotizacion(
            id=uuid.uuid1(),
            nombre=nuevo_nombre,
            proyecto_asociado=cotizacion_padre.proyecto_asociado,
            orden_compra=True,
            proveedor_asociado=cotizacion_padre.proveedor_asociado,
            usuario_modificacion=cotizacion_padre.usuario_modificacion,
        )
        nueva_cotizacion.save()
        for i in cotizacion_padre.contacto_asociado.all():
            nueva_cotizacion.contacto_asociado.add(i)
            nueva_cotizacion.save()
        if cotizacion_padre.fecha_salida:
            nueva_cotizacion.fecha_salida = cotizacion_padre.fecha_salida
        if cotizacion_padre.fecha_respuesta:
            nueva_cotizacion.fecha_respuesta = cotizacion_padre.fecha_respuesta
        if cotizacion_padre.fecha_actualizacion_precio:
            nueva_cotizacion.fecha_actualizacion_precio = (
                cotizacion_padre.fecha_actualizacion_precio
            )
        nueva_cotizacion.save()
        for i in productos:
            producto_proyecto = Producto_proyecto.objects.get(id=i)
            producto = producto_proyecto.proyecto
            precios_producto = producto.lista_precios.all().order_by("fecha")
            ultimo_precio = list(precios_producto).pop()
            cantidad = request.POST[i]
            nuevo_producto_proyecto_proyecto = Producto_proyecto_cantidades(
                id=uuid.uuid1(),
                proyecto_asociado_cantidades=cotizacion_padre.proyecto_asociado,
                precio=ultimo_precio,
                producto_asociado_cantidades=producto_proyecto,
                cantidades=cantidad,
            )
            nuevo_producto_proyecto_proyecto.save()
            nueva_cotizacion.productos_asociados.add(producto)
            nueva_cotizacion.productos_proyecto_asociados.add(
                nuevo_producto_proyecto_proyecto
            )
            nueva_cotizacion.save()
        if RMC.objects.filter(rut=factura).exists():
            destino_factura = RMC.objects.filter(rut=factura)[0]
        else:
            if factura == "76021113-3":
                nuevo_RMC = RMC(
                    rut="76021113-3",
                    nombre="INGENIERÍA Y SERVICIOS RMC LIMITADA",
                    giro="Servicios de Ingeniería",
                    direccion="Pasaje Las Cinerarias 550, Concón - Viña del Mar",
                )
            elif factura == "77241485-4":
                nuevo_RMC = RMC(
                    rut="76021113-3",
                    nombre="RMC INDUSTRIAL SPA",
                    giro="Servicios de Ingeniería",
                    direccion="Pasaje Las Cinerarias 550, Concón - Viña del Mar",
                )
            elif factura == "77241488-9":
                nuevo_RMC = RMC(
                    rut="77241488-9",
                    nombre="RMC EQUIPMENTS SPA",
                    giro="Servicios de Ingeniería",
                    direccion="Pasaje Las Cinerarias 550, Concón - Viña del Mar",
                )
            elif factura == "77289504-6":
                nuevo_RMC = RMC(
                    rut="77289504-6",
                    nombre="RMC CORPORATE SPA",
                    giro="Servicios de Ingeniería",
                    direccion="Pasaje Las Cinerarias 550, Concón - Viña del Mar",
                )
            elif factura == "77077958-8":
                nuevo_RMC = RMC(
                    rut="77077958-8",
                    nombre="RMC LABS SPA",
                    giro="Servicios de Ingeniería",
                    direccion="Pasaje Las Cinerarias 550, Concón - Viña del Mar",
                )
            else:
                nuevo_RMC = RMC(
                    rut="76021113-3",
                    nombre="INGENIERÍA Y SERVICIOS RMC LIMITADA",
                    giro="Servicios de Ingeniería",
                    direccion="Pasaje Las Cinerarias 550, Concón - Viña del Mar",
                )
            nuevo_RMC.save()
            destino_factura = nuevo_RMC
        if observaciones:
            nueva_orden_compra = Orden_compra(
                id=id_orden,
                cotizacion_padre=cotizacion_padre,
                cotizacion_hija=nueva_cotizacion,
                condicion_entrega=condicion_entrega,
                condiciones_pago=condicion_pago,
                forma_pago=forma_pago,
                destino_factura=destino_factura,
                observaciones=observaciones,
                fecha_envio=fecha_envio,
            )
        else:
            nueva_orden_compra = Orden_compra(
                id=id_orden,
                cotizacion_padre=cotizacion_padre,
                cotizacion_hija=nueva_cotizacion,
                condicion_entrega=condicion_entrega,
                condiciones_pago=condicion_pago,
                forma_pago=forma_pago,
                destino_factura=destino_factura,
                fecha_envio=fecha_envio,
            )
        nueva_orden_compra.save()
        planificadores = get_user_model().objects.filter(groups__name="Planificador")
        for i in planificadores:
            usuario_planificador = get_user_model().objects.get(correo=i.email)
            if not usuario_planificador.orden_compra:
                usuario_planificador.orden_compra = 0
                usuario_planificador.save()
            usuario_planificador.orden_compra -= 1
            usuario_planificador.save()
        usuario_cotizador = get_user_model().objects.get(correo=request.user.email)
        if not usuario_cotizador.orden_compra:
            usuario_cotizador.orden_compra = 0
            usuario_cotizador.save()
        usuario_cotizador.orden_compra -= 1
        usuario_cotizador.save()
        excel_oc(nueva_orden_compra.id)
        return redirect("/proyectos/mostrar_cotizacion/{}".format(nueva_cotizacion.id))


def editar_orden(request, id):
    orden_compra = Orden_compra.objects.get(id=id)
    if request.method == "GET":
        productos = orden_compra.cotizacion_hija.productos_proyecto_asociados.all()
        return render(
            request,
            "orden_compra/editar_orden.html",
            {"orden_compra": orden_compra, "productos": productos},
        )
    else:
        condicion_entrega = request.POST["condicion_entrega"]
        condicion_pago = request.POST["condicion_pago"]
        forma_pago = request.POST["forma_pago"]
        empresa = request.POST["empresa"]
        destino_factura = RMC.objects.get(rut=empresa)
        observaciones = request.POST["observaciones"]
        productos = request.POST.getlist("id_producto")
        orden_compra.condicion_entrega = condicion_entrega
        orden_compra.condiciones_pago = condicion_pago
        orden_compra.forma_pago = forma_pago
        orden_compra.destino_factura = destino_factura
        orden_compra.observaciones = observaciones
        fecha_envio = date.today()
        orden_compra.fecha_envio = fecha_envio
        orden_compra.save()
        cotizacion_hija = Cotizacion.objects.get(id=orden_compra.cotizacion_hija.id)
        for i in cotizacion_hija.productos_proyecto_asociados.all():
            cantidades = request.POST[i.id]
            producto_asociado = cotizacion_hija.productos_proyecto_asociados.filter(
                id=i.id
            )[0]
            for n in productos:
                if producto_asociado.id == n:
                    producto_asociado.cantidades = cantidades
                    producto_asociado.save()
        excel_oc(orden_compra.id)
        return redirect(
            "/proyectos/mostrar_cotizacion/{}".format(orden_compra.cotizacion_hija.id)
        )


def editar_status(request, id):
    orden_compra = Orden_compra.objects.get(id=id)
    if request.method == "GET":
        return render(
            request, "orden_compra/editar_status.html", {"orden_compra": orden_compra}
        )
    else:
        status_financiero = request.POST["status_financiero"]
        status_llegada = request.POST["status_llegada"]
        orden_compra.status_financiero = status_financiero
        orden_compra.status_llegada = status_llegada
        orden_compra.save()
        return redirect(
            "/proyectos/mostrar_cotizacion/{}".format(orden_compra.cotizacion_hija.id)
        )


def fecha_respuesta_editar_precio(cotizaciones):
    lista = []
    for i in cotizaciones:
        if i.orden_compra is not True:
            if i.fecha_respuesta:
                diccionario = {}
                diferencia_respuesta_salida = (
                    i.fecha_actualizacion_precio - i.fecha_respuesta
                )
                if diferencia_respuesta_salida.days < 0:
                    diferencia_respuesta_salida = diferencia_respuesta_salida * -1
                diccionario["category"] = i.nombre
                diccionario["amount"] = diferencia_respuesta_salida.days
                lista.append(diccionario)
    return json.dumps(lista)


def fecha_respuesta_cotizacion(cotizaciones):
    lista = []
    for i in cotizaciones:
        if i.orden_compra is not True:
            if i.fecha_respuesta:
                diccionario = {}
                diferencia_respuesta_cotizacion = i.fecha_respuesta - i.fecha_salida
                if diferencia_respuesta_cotizacion.days < 0:
                    diferencia_respuesta_cotizacion = (
                        diferencia_respuesta_cotizacion * -1
                    )
                diccionario["category"] = i.nombre
                diccionario["amount"] = diferencia_respuesta_cotizacion.days
                lista.append(diccionario)
    return json.dumps(lista)


def fecha_envio_orden(cotizaciones):
    lista = []
    for i in cotizaciones:
        diccionario = {}
        ordenes_compra = Orden_compra.objects.filter(cotizacion_hija=i)
        for n in ordenes_compra:
            if i.fecha_actualizacion_precio:
                diferencia_respuesta_orden = (
                    n.fecha_envio - i.fecha_actualizacion_precio
                ).days
            else:
                diferencia_respuesta_orden = 0
        diccionario["category"] = n.id
        diccionario["amount"] = diferencia_respuesta_orden
        lista.append(diccionario)
    return json.dumps(lista)


def graficos_proveedores(cotizaciones):
    diccionario_proveedores = {}
    for i in cotizaciones:
        if i.proveedor_asociado.nombre in diccionario_proveedores.keys():
            for n in i.productos_proyecto_asociados.all():
                if n.precio.tipo_cambio != "CLP":
                    diccionario_proveedores[i.proveedor_asociado.nombre] += int(
                        n.precio.valor
                        * n.precio.valor_cambio
                        * int(n.cantidades.split(".")[0])
                    )
                else:
                    diccionario_proveedores[i.proveedor_asociado.nombre] += int(
                        n.precio.valor * int(n.cantidades.split(".")[0])
                    )
        else:
            contador = 0
            for n in i.productos_proyecto_asociados.all():
                if contador == 0:
                    if n.precio.tipo_cambio != "CLP":
                        diccionario_proveedores[i.proveedor_asociado.nombre] = int(
                            n.precio.valor
                            * n.precio.valor_cambio
                            * int(n.cantidades.split(".")[0])
                        )
                    else:
                        diccionario_proveedores[i.proveedor_asociado.nombre] = int(
                            n.precio.valor * int(n.cantidades.split(".")[0])
                        )
                else:
                    if n.precio.tipo_cambio != "CLP":
                        diccionario_proveedores[i.proveedor_asociado.nombre] += int(
                            n.precio.valor
                            * n.precio.valor_cambio
                            * int(n.cantidades.split(".")[0])
                        )
                    else:
                        diccionario_proveedores[i.proveedor_asociado.nombre] += int(
                            n.precio.valor * int(n.cantidades.split(".")[0])
                        )
                contador += 1
    lista_proveedores = []
    for proveedor in diccionario_proveedores.keys():
        diccionario_aux = {}
        diccionario_aux["category"] = proveedor
        diccionario_aux["amount"] = diccionario_proveedores[proveedor]
        lista_proveedores.append(diccionario_aux)
    return json.dumps(lista_proveedores)


def graficos_clase(cotizaciones, proyecto, gastos_generales):
    diccionario_clase = {}
    diccionario_subclase = {}
    for i in cotizaciones:
        for n in i.productos_proyecto_asociados.all():
            subclase_modelo = (
                n.producto_asociado_cantidades.proyecto.subclase_set.all()[0]
            )
            subclase = subclase_modelo.nombre
            clase = subclase_modelo.clase_set.all()[0].nombre
            if clase in diccionario_clase.keys():
                if n.precio.tipo_cambio != "CLP":
                    diccionario_clase[clase] += int(
                        n.precio.valor
                        * n.precio.valor_cambio
                        * int(n.cantidades.split(".")[0])
                    )
                else:
                    diccionario_clase[clase] += int(
                        n.precio.valor * int(n.cantidades.split(".")[0])
                    )
            else:
                if n.precio.tipo_cambio != "CLP":
                    diccionario_clase[clase] = int(
                        n.precio.valor
                        * n.precio.valor_cambio
                        * int(n.cantidades.split(".")[0])
                    )
                else:
                    diccionario_clase[clase] = int(
                        n.precio.valor * int(n.cantidades.split(".")[0])
                    )
            if subclase in diccionario_subclase.keys():
                if n.precio.tipo_cambio != "CLP":
                    diccionario_subclase[subclase] += int(
                        n.precio.valor
                        * n.precio.valor_cambio
                        * int(n.cantidades.split(".")[0])
                    )
                    if n.precio.nombre_importacion:
                        diccionario_subclase[subclase] += int(
                            n.precio.valor_importación
                        ) * int(n.cantidades.split(".")[0])
                else:
                    diccionario_subclase[subclase] += int(
                        n.precio.valor * int(n.cantidades.split(".")[0])
                    )
                    if n.precio.nombre_importacion:
                        diccionario_subclase[subclase] += int(
                            n.precio.valor_importación
                        ) * int(n.cantidades.split(".")[0])
            else:
                if n.precio.tipo_cambio != "CLP":
                    diccionario_subclase[subclase] = int(
                        n.precio.valor
                        * n.precio.valor_cambio
                        * int(n.cantidades.split(".")[0])
                    )
                    if n.precio.nombre_importacion:
                        diccionario_subclase[subclase] += int(
                            n.precio.valor_importación
                        ) * int(n.cantidades.split(".")[0])
                else:
                    diccionario_subclase[subclase] = int(
                        n.precio.valor * int(n.cantidades.split(".")[0])
                    )
                    if n.precio.nombre_importacion:
                        diccionario_subclase[subclase] += int(
                            n.precio.valor_importación
                        ) * int(n.cantidades.split(".")[0])
    lista_subclase = []
    lista_presupuestos = []
    suma_costos = gastos_generales
    diccionario_ppto_subclases = {}
    for i in proyecto.presupuesto_subclases.all():
        diccionario_ppto_subclases[i.subclase.nombre] = i.valor
    for subclase in diccionario_subclase.keys():
        diccionario_aux = {}
        diccionario_aux_subclase = {}
        diccionario_aux_subclase_ppto = {}
        diccionario_aux["category"] = subclase
        diccionario_aux["amount"] = int(diccionario_subclase[subclase])
        suma_costos += diccionario_subclase[subclase]
        diccionario_aux_subclase["category"] = subclase
        diccionario_aux_subclase["position"] = 1
        diccionario_aux_subclase["value"] = int(diccionario_subclase[subclase])
        lista_presupuestos.append(diccionario_aux_subclase)
        diccionario_aux_subclase_ppto["category"] = subclase
        diccionario_aux_subclase_ppto["position"] = 0
        if subclase in diccionario_ppto_subclases.keys():
            diccionario_aux_subclase_ppto["value"] = int(
                diccionario_ppto_subclases[subclase]
            )
        else:
            diccionario_aux_subclase_ppto["value"] = 0
        lista_presupuestos.append(diccionario_aux_subclase_ppto)

        lista_subclase.append(diccionario_aux)
    lista_clase = []
    lista_presupuesto_total = []
    diccionario_aux_ppto = {}
    diccionario_aux_ppto2 = {}
    diccionario_aux_ppto["category"] = "Presupuesto"
    diccionario_aux_ppto["position"] = 0
    diccionario_aux_ppto["value"] = int(proyecto.presupuesto_total)
    diccionario_aux_ppto2["category"] = "Presupuesto"
    diccionario_aux_ppto2["position"] = 1
    diccionario_aux_ppto2["value"] = int(suma_costos)
    lista_presupuesto_total.append(diccionario_aux_ppto2)
    lista_presupuesto_total.append(diccionario_aux_ppto)
    for clase in diccionario_clase.keys():
        diccionario_aux = {}
        diccionario_aux["category"] = clase
        diccionario_aux["amount"] = int(diccionario_clase[clase])
        lista_clase.append(diccionario_aux)
    porcentaje_ppto_total = 0

    for i in lista_presupuesto_total:
        if i["position"] == 0:
            abajo = i["value"]
        elif i["position"] == 1:
            arriba = int(i["value"])
    porcentaje_ppto_total = arriba * 100 / abajo
    return [
        json.dumps(lista_clase),
        json.dumps(lista_subclase),
        json.dumps(lista_presupuestos),
        json.dumps(lista_presupuesto_total),
        porcentaje_ppto_total,
    ]


def graficos_gastos_generales(proyecto):
    lista_relacion_persona = []
    lista_gastos = []
    diccionario_relacion_persona = {}
    diccionario_gastos_generales = {}
    relaciones = proyecto.relacion_gastos.all()
    for i in relaciones:
        if i.rut_solicitante in diccionario_relacion_persona.keys():
            diccionario_relacion_persona[i.rut_solicitante] += 1
        else:
            diccionario_relacion_persona[i.rut_solicitante] = 1
        for n in i.gastos_generales.all():
            if n.razon_social in diccionario_gastos_generales.keys():
                diccionario_gastos_generales[n.razon_social] += n.monto
            else:
                diccionario_gastos_generales[n.razon_social] = n.monto
    for persona in diccionario_relacion_persona.keys():
        diccionario_aux = {}
        diccionario_aux["category"] = persona
        diccionario_aux["amount"] = int(diccionario_relacion_persona[persona])
        lista_relacion_persona.append(diccionario_aux)
    for razon_social in diccionario_gastos_generales.keys():
        diccionario_aux = {}
        diccionario_aux["category"] = razon_social
        diccionario_aux["position"] = 0
        diccionario_aux["value"] = int(diccionario_gastos_generales[razon_social])
        lista_gastos.append(diccionario_aux)
    return [json.dumps(lista_relacion_persona), json.dumps(lista_gastos)]


def graficos_importacion(cotizaciones):
    diccionario_importaciones = {}
    for i in cotizaciones:
        for x in i.productos_proyecto_asociados.all():
            if x.precio.nombre_importacion:
                importacion = Importaciones.objects.get(
                    codigo=x.precio.nombre_importacion
                )
                if importacion.codigo_referencial in diccionario_importaciones.keys():
                    diccionario_importaciones[importacion.codigo_referencial] += float(
                        x.precio.valor_importación
                    ) * float(x.cantidades)
                else:
                    diccionario_importaciones[importacion.codigo_referencial] = float(
                        x.precio.valor_importación
                    ) * float(x.cantidades)
    lista_importaciones = []
    for importacion in diccionario_importaciones.keys():
        diccionario_aux = {}
        diccionario_aux["category"] = importacion
        diccionario_aux["amount"] = int(diccionario_importaciones[importacion])
        lista_importaciones.append(diccionario_aux)
    return json.dumps(lista_importaciones)


def info_gasto(request, id):
    proyecto = Proyecto.objects.get(id=id)
    if request.method == "POST":
        return redirect("/proyectos/proyecto/{}".format(id))
    else:
        gastos_orden_compra = 0
        gastos_importacion = 0
        iva_orden_compra = 0
        # STATUS FINANCIERO
        no_pagado = 0
        cheque_a_fecha = 0
        en_proceso = 0
        pagado = 0
        cotizaciones_totales = Cotizacion.objects.filter(proyecto_asociado=proyecto)
        cotizaciones = Cotizacion.objects.filter(
            proyecto_asociado=proyecto, orden_compra=True
        )
        for i in cotizaciones:
            if i.orden_compra is True:
                orden_compra = Orden_compra.objects.filter(cotizacion_hija=i)
                if orden_compra[0].status_financiero:
                    if orden_compra[0].status_financiero == "CHEQUE A FECHA":
                        cheque_a_fecha += 1
                    elif orden_compra[0].status_financiero == "EN PROCESO":
                        en_proceso += 1
                    elif orden_compra[0].status_financiero == "PAGADO":
                        pagado += 1
                else:
                    no_pagado += 1
                for n in i.productos_proyecto_asociados.all():
                    if n.precio.tipo_cambio != "CLP":
                        precio_final = (
                            int(n.precio.valor_cambio)
                            * int(n.precio.valor)
                            * int(n.cantidades.split(".")[0])
                        )
                    else:
                        precio_final = int(n.precio.valor) * int(
                            n.cantidades.split(".")[0]
                        )
                    gastos_orden_compra += precio_final
                    iva_orden_compra += precio_final * 0.19
            for x in i.productos_proyecto_asociados.all():
                if x.precio.valor_importación:
                    gastos_importacion += int(x.precio.valor_importación) * int(
                        x.cantidades.split(".")[0]
                    )
        gastos_generales = 0
        iva_generales = 0
        for i in proyecto.relacion_gastos.all():
            gastos_generales += int(i.total_boleta)
            gastos_generales += int(i.total_factura)
            iva_generales = int(i.total_factura) * 0.19
        proveedores = graficos_proveedores(cotizaciones)
        clase = graficos_clase(cotizaciones, proyecto, gastos_generales)[0]
        subclase = graficos_clase(cotizaciones, proyecto, gastos_generales)[1]
        subclases_ppto = graficos_clase(cotizaciones, proyecto, gastos_generales)[2]
        ppto_total = graficos_clase(cotizaciones, proyecto, gastos_generales)[3]
        porcentaje_ppto_total = (
            str(graficos_clase(cotizaciones, proyecto, gastos_generales)[4]) + "%"
        )
        fecha_FCEP = fecha_respuesta_editar_precio(cotizaciones_totales)
        fecha_FRC = fecha_respuesta_cotizacion(cotizaciones_totales)
        fecha_EO = fecha_envio_orden(cotizaciones)
        gastos = [
            gastos_orden_compra,
            gastos_generales,
            gastos_orden_compra + gastos_generales + gastos_importacion,
            gastos_importacion,
        ]
        iva = [iva_orden_compra, iva_generales]
        status_financiero = [no_pagado, cheque_a_fecha, en_proceso, pagado]
        grafico_gastos_generales = graficos_gastos_generales(proyecto)[0]
        grafico_gastos_generales2 = graficos_gastos_generales(proyecto)[1]
        importaciones = graficos_importacion(cotizaciones)
        return render(
            request,
            "orden_compra/info_gasto.html",
            {
                "Proyecto": proyecto,
                "gastos": gastos,
                "IVA": iva,
                "status_financiero": status_financiero,
                "fecha_FCEP": fecha_FCEP,
                "fecha_FRC": fecha_FRC,
                "fecha_EO": fecha_EO,
                "proveedores": proveedores,
                "clase": clase,
                "subclase": subclase,
                "subclases_ppto": subclases_ppto,
                "ppto_total": ppto_total,
                "porcentaje_ppto_total": json.dumps(porcentaje_ppto_total),
                "grafico_gastos_generales": grafico_gastos_generales,
                "grafico_gastos_generales2": grafico_gastos_generales2,
                "importaciones": importaciones,
            },
        )
