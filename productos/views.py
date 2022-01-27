from django.shortcuts import render, redirect
from planificador.models import (
    Producto,
    Clase,
    SubClase,
    Precio,
    Filtro_producto,
    Producto_proveedor,
    Proveedor,
    Correlativo_producto,
    ImagenProducto,
)
from planificador.filters import Filtro_productoFilter
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django import forms
import openpyxl
import uuid
from planificador.api import api_token, get_locations
from planificador.notificaciones import crear_notificacion

UNIDAD_CHOICES = [
    ("", ""),
    ("Unidad", "Unidad"),
    ("Kilos", "Kilos"),
    ("Metros", "Metros"),
    ("Litros", "Litros"),
    ("Días", "Días"),
    ("Horas", "Horas"),
    ("Par", "Par"),
    ("Bolsas", "Bolsas"),
]


class ImageForm(forms.ModelForm):
    unidad = forms.CharField(
        required=False,
        widget=forms.Select(attrs={"class": "custom-select"}, choices=UNIDAD_CHOICES),
    )
    kilos = forms.CharField(required=False)
    imagen = forms.ImageField(required=False)

    class Meta:
        model = Producto
        fields = ("unidad", "kilos", "imagen")


def isfloat(value):
    try:
        float(value)
        return True
    except ValueError:
        return False


# Mostrar productos
@login_required(login_url="/login")
def productos(request):
    productos = Producto.objects.all()
    lista_productos = []
    for i in productos:
        subclase = i.subclase_set.all()[0]
        clase = subclase.clase_set.all()[0]
        lista_productos.append([i, subclase, clase])
    productos = Filtro_producto.objects.all()
    myFilter = Filtro_productoFilter(request.GET, queryset=productos)
    payload = {
        "Productos": lista_productos,
        "myFilter": myFilter,
        "len": len(lista_productos),
    }
    return render(request, "productos/productos.html", payload)


def nuevo_producto_planilla(request):
    if request.method == "POST":
        datos_fallados = []
        booleano_fallados = False
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["producto"]
        contador_creado = 0
        creado = False
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            id = row_data[0].upper()
            nombre = row_data[1].upper()
            unidad = row_data[2].upper()
            kilos = row_data[3].upper()
            subclase = row_data[4].upper()
            if id != "ID":
                if id == "NONE" or nombre == "NONE":
                    if not (
                        id == "NONE"
                        and nombre == "NONE"
                        and unidad == "NONE"
                        and kilos == "NONE"
                        and subclase == "NONE"
                    ):
                        aux = []
                        aux.append(row_data[0])
                        aux.append(row_data[1])
                        aux.append("No se ingresó ID o Nombre")
                        datos_fallados.append(aux)
                else:
                    fecha_actualizacion = datetime.now()
                    if Producto.objects.filter(id=id).exists():
                        aux = []
                        aux.append(row_data[0])
                        aux.append(row_data[1])
                        aux.append("Producto con id:{} ya existe".format(id))
                        datos_fallados.append(aux)
                    else:
                        if not SubClase.objects.filter(nombre=subclase).exists():
                            aux = []
                            aux.append(row_data[0])
                            aux.append(row_data[1])
                            aux.append("La SubClase {} no existe".format(subclase))
                            datos_fallados.append(aux)
                        else:
                            nuevo_producto = Producto(
                                id=id,
                                nombre=nombre,
                                fecha_actualizacion=fecha_actualizacion,
                            )
                            if unidad != "None":
                                nuevo_producto.unidad = unidad
                            if kilos != "None":
                                es_float = isfloat(kilos)
                                if es_float:
                                    nuevo_producto.kilos = kilos
                                else:
                                    aux = []
                                    aux.append(row_data[0])
                                    aux.append(row_data[1])
                                    aux.append(
                                        "Producto creado sin kilos. No es un número"
                                    )
                                    datos_fallados.append(aux)
                            #NUEVOS
                            if row_data[11] != "None":
                                nuevo_producto.superficie = row_data[11]
                            if row_data[12] != "None":
                                nuevo_producto.perimetro = row_data[12]
                            if row_data[13] != "None":
                                nuevo_producto.hh1 = row_data[13]
                            if row_data[14] != "None":
                                nuevo_producto.hh2 = row_data[14]
                            if row_data[15] != "None":
                                nuevo_producto.hh3 = row_data[15]
                            if row_data[16] != "None":
                                nuevo_producto.hh4 = row_data[16]
                            if row_data[17] != "None":
                                nuevo_producto.hh5 = row_data[17]
                            if row_data[18] != "None":
                                nuevo_producto.hh6 = row_data[18]
                            if row_data[19] != "None":
                                nuevo_producto.hh7 = row_data[19]
                            if row_data[20] != "None":
                                nuevo_producto.hh8 = row_data[20]
                            if row_data[21] != "None":
                                nuevo_producto.d1 = row_data[21]
                            if row_data[22] != "None":
                                nuevo_producto.d2 = row_data[22]
                            nuevo_producto.save()
                            # CREAMOS UN PRECIO (PARA POBLAR BBDD)
                            if (
                                row_data[5] != "None"
                                and row_data[6] != "None"
                                and row_data[10] != "None"
                            ):
                                nuevo_precio = Precio(
                                    id=uuid.uuid1(),
                                    fecha=row_data[5],
                                    valor=row_data[6],
                                    nombre_proveedor=row_data[10].upper(),
                                )
                                nuevo_precio.save()
                                if row_data[7] != "None":
                                    nuevo_precio.valor_importación = row_data[7]
                                if row_data[8] != "None":
                                    nuevo_precio.tipo_cambio = row_data[8].upper()
                                if row_data[9] != "None":
                                    nuevo_precio.valor_cambio = row_data[9]
                                nuevo_precio.save()
                                nuevo_producto.lista_precios.add(nuevo_precio)
                            contador_creado += 1
                            creado = True
                            sub_clase = SubClase.objects.get(nombre=subclase)
                            sub_clase.productos.add(nuevo_producto)
                            clase = sub_clase.clase_set.all()
                            nuevo_filtro_producto = Filtro_producto(
                                nombre_producto=nombre,
                                nombre_clase=clase[0].nombre,
                                id_producto=id,
                                nombre_subclase=subclase,
                            )
                            nuevo_filtro_producto.save()
                            sub_clase.save()
        if len(datos_fallados) != 0:
            booleano_fallados = True
        return render(
            request,
            "productos/resultado_planilla_productos.html",
            {"Fallo": datos_fallados, "Booleano": booleano_fallados},
        )
    else:
        return render(request, "productos/nuevo_producto_planilla.html")


def nuevo_producto_interno_planilla(request):
    if request.method == "POST":
        datos_fallados = []
        booleano_fallados = False
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["producto_interno"]
        contador_creado = 0
        creado = False
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            id = row_data[0].upper()
            nombre = row_data[1].upper()
            unidad = row_data[2].upper()
            proveedor = row_data[3].upper()
            valor = row_data[4]
            moneda = row_data[5].upper()
            valor_moneda = row_data[6]
            subclase = row_data[7].upper()
            if id != "ID":
                if id == "NONE" or nombre == "NONE":
                    if not (id == "NONE" and nombre == "NONE"):
                        aux = []
                        aux.append(row_data[0])
                        aux.append(row_data[1])
                        aux.append("No se ingresó ID o Nombre")
                        datos_fallados.append(aux)
                else:
                    fecha_actualizacion = datetime.now()
                    if Producto.objects.filter(id=id).exists():
                        aux = []
                        aux.append(row_data[0])
                        aux.append(row_data[1])
                        aux.append("Producto con id:{} ya existe".format(id))
                        datos_fallados.append(aux)
                    else:
                        if not SubClase.objects.filter(nombre=subclase).exists():
                            if subclase != "SUBCLASE":
                                aux = []
                                aux.append(row_data[0])
                                aux.append(row_data[1])
                                aux.append("La SubClase {} no existe".format(subclase))
                                datos_fallados.append(aux)
                        else:
                            nuevo_producto = Producto(
                                id=id,
                                nombre=nombre,
                                fecha_actualizacion=fecha_actualizacion,
                            )
                            if unidad != "None":
                                nuevo_producto.unidad = unidad
                            nuevo_producto.proveedor_interno = proveedor
                            nuevo_producto.save()
                            if (
                                row_data[4] != "None"
                                and row_data[5] != "None"
                                and row_data[6] != "None"
                            ):
                                nuevo_precio = Precio(
                                    id=uuid.uuid1(),
                                    fecha=fecha_actualizacion,
                                    valor=valor,
                                    nombre_proveedor=proveedor,
                                )
                                nuevo_precio.save()
                                if row_data[5] != "None":
                                    nuevo_precio.tipo_cambio = moneda
                                if row_data[6] != "None":
                                    nuevo_precio.valor_cambio = valor_moneda
                                nuevo_precio.save()
                                nuevo_producto.lista_precios.add(nuevo_precio)
                            contador_creado += 1
                            creado = True
                            sub_clase = SubClase.objects.get(nombre=subclase)
                            sub_clase.productos.add(nuevo_producto)
                            clase = sub_clase.clase_set.all()
                            nuevo_filtro_producto = Filtro_producto(
                                nombre_producto=nombre,
                                nombre_clase=clase[0].nombre,
                                id_producto=id,
                                nombre_subclase=subclase,
                            )
                            nuevo_filtro_producto.save()
                            sub_clase.save()
        if len(datos_fallados) != 0:
            booleano_fallados = True
        return render(
            request,
            "productos/resultado_planilla_productos_interno.html",
            {"Fallo": datos_fallados, "Booleano": booleano_fallados},
        )
    else:
        return render(request, "productos/nuevo_producto_interno_planilla.html")


# Agregar producto
@login_required(login_url="/login")
def agregar_producto(request):
    clases = Clase.objects.all()
    subclases = SubClase.objects.all()
    lista_clases = [[i, [n for n in i.subclases.all()]] for i in clases]
    payload = {"Clases": clases, "Subclases": subclases, "lista_clases": lista_clases}
    return render(request, "productos/crear_producto.html", payload)


def recibir_datos_producto(request):
    get = request.GET
    if Correlativo_producto.objects.filter(producto=0).exists():
        correlativo = Correlativo_producto.objects.get(producto=0)
        correlativo.numero += 1
        correlativo.save()
    else:
        correlativo = Correlativo_producto(producto=0, numero=9000000)
        correlativo.save()
    if get["peso"] and (get["unidad"] != "ELEGIR UNIDAD" and get["unidad"]):
        nuevo_producto = Producto(
            id=correlativo.numero, nombre=get["nombre"], unidad=get["unidad"], kilos=get["peso"]
        )
        nuevo_producto.save()
    elif get["peso"] and (get["unidad"] == "ELEGIR UNIDAD"):
        nuevo_producto = Producto(id=correlativo.numero, nombre=get["nombre"], kilos=get["peso"])
        nuevo_producto.save()
    elif (get["unidad"] != "ELEGIR UNIDAD") and not get["peso"]:
        nuevo_producto = Producto(id=correlativo.numero, nombre=get["nombre"], unidad=get["unidad"])
        nuevo_producto.save()
    else:
        nuevo_producto = Producto(id=correlativo.numero, nombre=get["nombre"])
        nuevo_producto.save()
    if get["kilos"] and get["kilos"]!="None":
        nuevo_producto.kilos = get["kilos"]
    if get["superficie"] and get["superficie"]!="None":
        nuevo_producto.superficie = get["superficie"]
    if get["perimetro"] and get["perimetro"]!="None":
        nuevo_producto.perimetro = get["perimetro"]
    if get["hh1"] and get["hh1"]!="None":
        nuevo_producto.hh1 = get["hh1"]
    if get["hh2"] and get["hh2"]!="None":
        nuevo_producto.hh2 = get["hh2"]
    if get["hh3"] and get["hh3"]!="None":
        nuevo_producto.hh3 = get["hh3"]
    if get["hh4"] and get["hh4"]!="None":
        nuevo_producto.hh4 = get["hh4"]
    if get["hh5"] and get["hh5"]!="None":
        nuevo_producto.hh5 = get["hh5"]
    if get["hh6"] and get["hh6"]!="None":
        nuevo_producto.hh6 = get["hh6"]
    if get["hh7"] and get["hh7"]!="None":
        nuevo_producto.hh7 = get["hh7"]
    if get["hh8"] and get["hh8"]!="None":
        nuevo_producto.hh8 = get["hh8"]
    if get["d1"] and get["d1"]!="None":
        nuevo_producto.d1 = get["d1"]
    if get["d2"] and get["d2"]!="None":
        nuevo_producto.d2 = get["d2"]
    nuevo_producto.save()
    subclase = SubClase.objects.get(nombre=get["subclase"])
    subclase.productos.add(nuevo_producto)
    clase = subclase.clase_set.all()
    nuevo_filtro_producto = Filtro_producto(
        nombre_producto=get["nombre"],
        nombre_clase=clase[0].nombre,
        id_producto=correlativo.numero,
        nombre_subclase=subclase.nombre,
    )
    nuevo_filtro_producto.save()
    return redirect("/productos/producto/{}".format(nuevo_producto.id))


# Vista producto
@login_required(login_url="/login")
def producto(request, id):
    producto = Producto.objects.get(id=id)
    lista_precios = producto.lista_precios.all()
    a = lista_precios.order_by("-fecha")
    sub_clase = producto.subclase_set.all()[0]
    clase = sub_clase.clase_set.all()[0]
    imagenes = producto.imagen.all()
    if Producto_proveedor.objects.filter(proyecto=producto).exists():
        nombre_proveedor = Producto_proveedor.objects.filter(proyecto=producto)
    else:
        nombre_proveedor = ""
    payload = {
        "Producto": producto,
        "lista_precios": a,
        "Subclase": sub_clase,
        "Clase": clase,
        "nombre_proveedor": nombre_proveedor,
        "imagenes": imagenes,
    }
    return render(request, "productos/producto.html", payload)


@login_required(login_url="/login")
def nuevo_proveedor_producto(request):
    if request.method == "POST":
        datos_fallados = []
        booleano_fallados = False
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["producto_proveedor"]
        contador_creado = 0
        creado = False
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            nombre_producto = row_data[0].upper()
            proveedor = row_data[1].upper()
            nombre_producto_proveedor = row_data[2].upper()
            if nombre_producto != "NOMBRE_PRODUCTO_RMC":
                if (
                    nombre_producto == "NONE"
                    or proveedor == "NONE"
                    or nombre_producto_proveedor == "NONE"
                ):
                    if not (
                        nombre_producto == "NONE"
                        and proveedor == "NONE"
                        and nombre_producto_proveedor == "NONE"
                    ):
                        datos_fallados.append([row_data[0], row_data[1], row_data[2],
                                              "No se ingresó o nombre producto RMC o nombre del proveedor o nombre del producto para proveedor"])
                else:
                    if not Producto_proveedor.objects.filter(
                        nombre_RMC=nombre_producto,
                        nombre_proveedor=nombre_producto_proveedor,
                    ).exists():
                        if not Producto.objects.filter(nombre=nombre_producto).exists():
                            datos_fallados.append([row_data[0], row_data[1], row_data[2],
                                                  "Producto con nombre:{} no existe".format(
                                                  nombre_producto)])
                        elif not Proveedor.objects.filter(nombre=proveedor).exists():
                            datos_fallados.append([row_data[0], row_data[1], row_data[2],
                                                  "Proveedor con nombre:{} no existe".format(proveedor)])
                        else:
                            proveedor_ingreso = Proveedor.objects.get(nombre=proveedor)
                            producto = Producto.objects.get(nombre=nombre_producto)
                            if not Producto_proveedor.objects.filter(
                                producto=proveedor_ingreso, proyecto=producto
                            ).exists():
                                nuevo_producto_proveedor = Producto_proveedor(
                                    producto=proveedor_ingreso,
                                    proyecto=producto,
                                    nombre_RMC=nombre_producto,
                                    nombre_proveedor=nombre_producto_proveedor,
                                )
                                nuevo_producto_proveedor.save()
                                creado = True
                                contador_creado += 1
                            else:
                                producto_proveedor = Producto_proveedor.objects.get(
                                    producto=proveedor_ingreso, proyecto=producto
                                )
                                producto_proveedor.nombre_proveedor = (
                                    nombre_producto_proveedor
                                )
                                producto_proveedor.save()
                                creado = True
                                contador_creado += 1
                    else:
                        datos_fallados.append(row_data[0], row_data[1], row_data[2], "Ya existe el mismo nombre en relación")
        if len(datos_fallados) != 0:
            booleano_fallados = True
        payload = {"Fallo": datos_fallados, "Booleano": booleano_fallados}
        return render(request, "productos/resultado_planilla_proveedor_productos.html", payload)
    else:
        return render(request, "productos/nuevo_proveedor_producto.html")


# Edición producto
@login_required(login_url="/login")
def mostrar_edicion_producto(request, id):
    producto = Producto.objects.get(id=id)
    if request.method == "POST":
        post = request.POST
        imagen = request.FILES["imagen"]
        if imagen:
            nueva_imagen = ImagenProducto(id=uuid.uuid1(), imagen=imagen)
            nueva_imagen.save()
        producto.unidad = post["unidad"]
        if post["kilos"] and post["kilos"]!="None":
            producto.kilos = post["kilos"]
        if post["superficie"] and post["superficie"]!="None":
            producto.superficie = post["superficie"]
        if post["perimetro"] and post["perimetro"]!="None":
            producto.perimetro = post["perimetro"]
        if post["hh1"] and post["hh1"]!="None":
            producto.hh1 = post["hh1"]
        if post["hh2"] and post["hh2"]!="None":
            producto.hh2 = post["hh2"]
        if post["hh3"] and post["hh3"]!="None":
            producto.hh3 = post["hh3"]
        if post["hh4"] and post["hh4"]!="None":
            producto.hh4 = post["hh4"]
        if post["hh5"] and post["hh5"]!="None":
            producto.hh5 = post["hh5"]
        if post["hh6"] and post["hh6"]!="None":
            producto.hh6 = post["hh6"]
        if post["hh7"] and post["hh7"]!="None":
            producto.hh7 = post["hh7"]
        if post["hh8"] and post["hh8"]!="None":
            producto.hh8 = post["hh8"]
        if post["d1"] and post["d1"]!="None":
            producto.d1 = post["d1"]
        if post["d2"] and post["d2"]!="None":
            producto.d2 = post["d2"]
        producto.imagen.add(nueva_imagen)
        producto.save()
        return redirect("/productos/producto/{}".format(producto.id))
    else:
        subclases = SubClase.objects.all()
        payload = {"Producto": producto, "Subclases": subclases}
        return render(request, "productos/editar_producto.html", payload)


# Eliminar producto
@login_required(login_url="/login")
def eliminar_producto(request, id):
    producto = Producto.objects.get(id=id)
    filtro = Filtro_producto.objects.get(nombre_producto=producto.nombre)
    producto.delete()
    filtro.delete()
    return redirect("/productos")


# Eliminar producto
@login_required(login_url="/login")
def mostrar_ubicaciones(request):
    access_token = api_token()
    locations_response = get_locations(access_token)
    payload = {"Ubicaciones": locations_response}
    return render(request, "productos/mostrar_ubicaciones.html", payload)
